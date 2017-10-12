"""
author: krozin@gmail.com. 2017.
"""

import datetime as dt
import json
import os

import BeautifulSoup as bs
import openpyxl

SPEC_USER = 'Jon Dou'
CUSTOMER = 'ibm'
PROJECT = 'R1/team1'
SPEC_STATUS = 'team1_done'

class Bugs2Excel(object):

    def __init__(self, workitems, teams):
        self.buglist = workitems
        self.teams = teams

    def get_team(self, name):
        for team, names in self.teams.items():
            if name in names:
                return team
        return 'Unknown'

    def save_bugs_to_file(self, filepath):
        if os.path.exists(filepath):
            wb = openpyxl.load_workbook(filename=filepath)
            # remove old data sheets
            if 'input' in wb:
                wb.remove(wb['input'])
            if 'input_history' in wb:
                wb.remove(wb['input_history'])
            if 'results' in wb:
                wb.remove(wb['results'])
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb['Sheet'])  # Remove default sheet

        wb.create_sheet("input")
        wb.create_sheet("input_history")
        wb.create_sheet("results")

        sheet = wb['input']
        sheet_history = wb['input_history']
        sheet_results = wb['results']

        sheet['A1'] = "itemId"
        sheet['B1'] = "stateId"
        sheet['C1'] = "locationUri"
        sheet['D1'] = 'modifiedBy'
        sheet['E1'] = "modified"
        sheet['F1'] = "contextId"
        sheet['G1'] = "projectArea"
        sheet['H1'] = "category"
        sheet['I1'] = "internalTags"
        sheet['J1'] = "internalPriority"
        sheet['K1'] = "creator"
        sheet['L1'] = "dueDate"
        sheet['M1'] = "duration"
        sheet['N1'] = "correctedEstimate"
        sheet['O1'] = "timeSpent"
        sheet['P1'] = "internalSequenceValue"
        sheet['Q1'] = "startDate"
        sheet['R1'] = "creationDate"
        sheet['S1'] = "resolver"
        sheet['T1'] = "resolutionDate"
        sheet['U1'] = "target"
        sheet['V1'] = "owner"
        sheet['W1'] = "description"
        sheet['X1'] = "summary"
        sheet['Y1'] = "internalState"
        sheet['Z1'] = "internalResolution"
        sheet['AA1'] = "internalSeverity"
        sheet['AB1'] = "foundIn"
        sheet['AC1'] = "workItemType"
        sheet['AD1'] = "id"
        sheet['AE1'] = "internalComments"
        sheet['AF1'] = "internalSubscriptions"
        sheet['AG1'] = "archived"
        sheet['AH1'] = "com.ibm.team.apt.attribute.planitem.priority"
        sheet['AI1'] = "com.ibm.team.apt.attribute.planitem.newRanking"
        sheet['AJ1'] = "com.{}.team.workitem.attribute.parenturi".format(CUSTOMER)
        sheet['AK1'] = "com.{}.team.workitem.attribute.resolutionnotes".format(CUSTOMER)
        sheet['AL1'] = "com.{}.team.workitem.attribute.rygstatus".format(CUSTOMER)
        sheet['AM1'] = "com.{}.team.workitem.attribute.rootcause".format(CUSTOMER)
        sheet['AN1'] = "com.{}.team.workitem.attribute.defect.issuescore".format(CUSTOMER)
        sheet['AO1'] = "com.{}.team.workitem.attribute.syncproxyattribute".format(CUSTOMER)
        sheet['AP1'] = "com.{}.team.workitem.attribute.modelyear".format(CUSTOMER)
        sheet['AQ1'] = "com.{}.team.workitem.attribute.suppliername".format(CUSTOMER)
        sheet['AR1'] = "com.{}.team.workitem.attribute.supplierreferenceid".format(CUSTOMER)
        sheet['AS1'] = ("com.{}.team.workitem.attribute."
                        "externalfeedbackstatus".format(CUSTOMER))
        sheet['AT1'] = ("com.{}.team.workitem.attribute."
                        "issueoccurrence_description".format(CUSTOMER))
        sheet['AU1'] = ("com.{}.team.workitem.attribute."
                        "issuedetection_description".format(CUSTOMER))
        sheet['AV1'] = ("com.{}.team.workitem.attribute."
                        "issueseverity_description".format(CUSTOMER))
        sheet['AW1'] = ("com.{}.team.workitem.attribute."
                        "testingstage".format(CUSTOMER))
        sheet['AX1'] = "com.{}.team.workitem.attribute.rootcausetype".format(CUSTOMER)
        sheet['AY1'] = "com.{}.team.workitem.attribute.defectdescription".format(CUSTOMER)
        sheet['AZ1'] = "com.{}.team.workitem.attribute.fixedin".format(CUSTOMER)
        sheet['BA1'] = "com.{}.team.workitem.attribute.testedin".format(CUSTOMER)
        sheet['BB1'] = "com.{}.team.workitem.attribute.actualINT".format(CUSTOMER)
        sheet['BC1'] = "com.{}.team.workitem.attribute.estimateINT".format(CUSTOMER)
        sheet['BD1'] = ("com.{}.team.workitem.attribute."
                        "aggregatedEstimate".format(CUSTOMER))
        sheet['BE1'] = "com.{}.team.workitem.attribute.aggregatedActual".format(CUSTOMER)
        sheet['BF1'] = "com.{}.team.workitem.attribute.rootcausehtml".format(CUSTOMER)
        sheet['BG1'] = "teamArea"
        sheet['BH1'] = "stateGroup"
        sheet['BI1'] = 'Children'
        sheet['BJ1'] = 'Blocks'
        sheet['BK1'] = 'Parents'
        sheet['BL1'] = 'lastComment'

        sheet_history['A1'] = "itemId"
        sheet_history['B1'] = "modifiedDate"
        sheet_history['C1'] = "userId"
        sheet_history['D1'] = 'content'
        sheet_history['E1'] = "parsed_history"
        sheet_history['F1'] = "owner"
        sheet_history['G1'] = "assigned_date"
        sheet_history['H1'] = "expiration_date"
        sheet_history['I1'] = "rca_date"
        sheet_history['J1'] = "closed_date"
        sheet_history['K1'] = "closed_week"

        sheet_results['A1'] = 'Id'
        sheet_results['B1'] = 'Title'
        sheet_results['C1'] = 'Priority'
        sheet_results['D1'] = 'Team'
        sheet_results['E1'] = 'Defect owner'
        sheet_results['F1'] = 'Domain'
        sheet_results['G1'] = 'Assigned Date'
        sheet_results['H1'] = 'Expiration Date'
        sheet_results['I1'] = 'Root Cause Defined Date'
        sheet_results['J1'] = 'Closed date'
        sheet_results['K1'] = 'Week out'
        sheet_results['L1'] = 'Week in'
        sheet_results['M1'] = 'RCA'
        sheet_results['N1'] = 'Status'
        sheet_results['O1'] = 'Tags'
        sheet_results['P1'] = 'Lead time'
        sheet_results['Q1'] = 'State'
        sheet_results['R1'] = 'Planned for'
        sheet_results['S1'] = 'Root Cause'

        p = 2
        for i, j in enumerate(self.buglist):
            n = i + 2
            it = j.get('soapenv:Envelope').get('soapenv:Body').get(
                'response').get('returnValue').get('value')

            sheet['A{}'.format(n)] = "{}".format(it.get('itemId'))
            sheet['B{}'.format(n)] = "{}".format(it.get('stateId'))
            sheet['C{}'.format(n)] = "{}".format(it.get('locationUri'))

            last_comment = ''
            for ij, a in enumerate(it.get('attributes')):

                if a.get('key') == 'modifiedBy':
                    sheet['D{}'.format(n)] = "{}".format(
                        a.get('value').get('userId'))

                elif a.get('key') == 'modified':
                    sheet['E{}'.format(n)] = "{}".format(
                        a.get('value').get('userId'))

                elif a.get('key') == 'contextId':
                    sheet['F{}'.format(n)] = "{}:{}".format(
                        a.get('value').get('id'), a.get('value').get('label'))
                elif a.get('key') == 'projectArea':
                    sheet['G{}'.format(n)] = "{}:{}".format(
                        a.get('value').get('id'), a.get('value').get('label'))
                elif a.get('key') == 'category':
                    sheet['H{}'.format(n)] = "{}:{}".format(
                        a.get('value').get('id'), a.get('value').get('label'))
                elif a.get('key') == 'internalTags':
                    val = a.get('value').get('content', '') or 'None'
                    sheet['I{}'.format(n)] = val.encode("utf8")

                elif a.get('key') == 'internalPriority':
                    sheet['J{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif a.get('key') == 'creator':
                    sheet['K{}'.format(n)] = "{}".format(
                        a.get('value').get('userId'))

                elif a.get('key') == 'dueDate':
                    sheet['L{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif a.get('key') == 'duration':
                    sheet['M{}'.format(n)] = "{}:{}".format(
                        a.get('value').get('id'), a.get('value').get('label'))

                elif a.get('key') == 'correctedEstimate':
                    sheet['N{}'.format(n)] = "{}:{}".format(
                        a.get('value').get('id'), a.get('value').get('label'))

                elif a.get('key') == 'timeSpent':
                    sheet['O{}'.format(n)] = "{}:{}".format(
                        a.get('value').get('id'), a.get('value').get('label'))

                elif a.get('key') == 'internalSequenceValue':
                    val = a.get('value').get('content') or 'None'
                    sheet['P{}'.format(n)] = val.encode("utf8")

                elif a.get('key') == 'startDate':
                    sheet['Q{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif a.get('key') == 'creationDate':
                    sheet['R{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif a.get('key') == 'resolver':
                    sheet['S{}'.format(n)] = "{}".format(
                        a.get('value').get('userId'))

                elif a.get('key') == 'resolutionDate':
                    sheet['T{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif a.get('key') == 'target':
                    sheet['U{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif a.get('key') == 'owner':
                    sheet['V{}'.format(n)] = "{}".format(
                        a.get('value').get('userId'))

                elif a.get('key') == 'description':
                    val = a.get('value').get('content', '') or 'None'
                    sheet['W{}'.format(n)] = val.encode("utf8")

                elif a.get('key') == 'summary':
                    val = a.get('value').get('content', 'None') or 'None'
                    sheet['X{}'.format(n)] = "{}".format(val.encode("utf8"))

                elif a.get('key') == 'internalState':
                    sheet['Y{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif a.get('key') == 'internalResolution':
                    sheet['Z{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif a.get('key') == 'internalSeverity':
                    sheet['AA{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif a.get('key') == 'foundIn':
                    sheet['AB{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif a.get('key') == 'workItemType':
                    sheet['AC{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif a.get('key') == 'id':
                    sheet['AD{}'.format(n)] = "{}".format(
                        a.get('value').get('id'))

                elif a.get('key') == 'internalComments':
                    if isinstance(a.get('value').get('items'), list):
                        comments = []
                        for ii in a.get('value').get('items'):
                            content = ii.get('content') or 'None'
                            comments.append(
                                "\n\nCREATOR:{} CREATED:{}\n{}".format(
                                    ii.get('creator').get('userId'),
                                    ii.get('creationDate'),
                                    content.encode("utf8")))
                        internalComments = "\n".join(comments)

                        ii = a.get('value').get('items')[-1]
                        val = ii.get('content', 'None') or 'None'
                        last_comment = "CREATOR:{} CREATED:{}"\
                                       "\n{}".format(
                                           ii.get('creator', {}).get(
                                               'userId'),
                                           ii.get('creationDate'),
                                           val.encode("utf8"))

                    elif isinstance(a.get('value').get('items'), dict):
                        val = a.get(
                            'value').get('items').get('content') or 'None'
                        internalComments = "CREATOR:{} CREATED:{}"\
                                           "\n{}".format(a.get('value').get(
                                               'items').get(
                                               'creator', {}).get('userId'),
                                               a.get('value').get('items').get(
                                                   'creationDate'),
                                               val.encode("utf8"))
                        last_comment = internalComments

                    sheet['AE{}'.format(n)] = "{}".format(internalComments)
                    sheet['BL{}'.format(n)] = "{}".format(last_comment)

                elif a.get('key') == 'internalSubscriptions':
                    if isinstance(a.get('value').get('items'), list):
                        internalSubscriptions = "\n".join(
                            [ii.get('userId')
                             for ii in a.get('value').get('items')])
                    if isinstance(a.get('value').get('items'), dict):
                        internalSubscriptions = a.get(
                            'value').get('items').get('userId')
                    sheet['AF{}'.format(n)] = "{}".format(
                        internalSubscriptions)

                elif a.get('key') == 'archived':
                    sheet['AG{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif ('com.ibm.team.apt.attribute.planitem.priority.' in
                        a.get('key')):
                    sheet['AH{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif ('com.ibm.team.apt.attribute.planitem.newRanking.' in
                        a.get('key')):
                    sheet['AI{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.parenturi'):
                    val = a.get('value').get('content') or 'None'
                    sheet['AJ{}'.format(n)] = val.encode("utf8")

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.resolutionnotes'):
                    val = a.get('value').get('content', '') or 'None'
                    sheet['AK{}'.format(n)] = val.encode("utf8")

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.rygstatus'):
                    sheet['AL{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.rootcause'):
                    val = a.get('value').get('content', '') or 'None'
                    sheet['AM{}'.format(n)] = val.encode("utf8")

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.defect.issuescore'):
                    sheet['AN{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.syncproxyattribute'):
                    val = a.get('value').get('content', '') or 'None'
                    sheet['AO{}'.format(n)] = val.encode("utf8")

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.modelyear'):
                    sheet['AP{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.suppliername'):
                    val = a.get('value').get('content', '') or 'None'
                    sheet['AQ{}'.format(n)] = val.encode("utf8")

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.'
                        'supplierreferenceid'):
                    val = a.get('value').get('content', '') or 'None'
                    sheet['AR{}'.format(n)] = val.encode("utf8")

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.'
                        'externalfeedbackstatus'):
                    sheet['AS{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.'
                        'issueoccurrence_description'):
                    sheet['AT{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.'
                        'issuedetection_description'):
                    sheet['AU{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.'
                        'issueseverity_description'):
                    sheet['AV{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.testingstage'):
                    sheet['AW{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.rootcausetype'):
                    sheet['AX{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.defectdescription'):
                    val = a.get('value').get('content', '') or 'None'
                    sheet['AY{}'.format(n)] = val.encode("utf8")

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.fixedin'):
                    sheet['AZ{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.testedin'):
                    sheet['BA{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.actualINT'):
                    sheet['BB{}'.format(n)] = "{}".format(
                        a.get('value').get('id'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.estimateINT'):
                    sheet['BC{}'.format(n)] = "{}".format(
                        a.get('value').get('id'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.'
                        'aggregatedEstimate'):
                    sheet['BD{}'.format(n)] = "{}".format(
                        a.get('value').get('id'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.'
                        'aggregatedActual'):
                    sheet['BE{}'.format(n)] = "{}".format(
                        a.get('value').get('id'))

                elif (a.get('key') ==
                        'com.{}.team.workitem.attribute.rootcausehtml'):
                    val = a.get('value').get('content', '') or 'None'
                    sheet['BF{}'.format(n)] = val.encode("utf8")

                elif a.get('key') == 'teamArea':
                    sheet['BG{}'.format(n)] = "{}".format(
                        a.get('value').get('label'))

                elif a.get('key') == 'stateGroup':
                    sheet['BH{}'.format(n)] = "{}".format(
                        a.get('value').get('id'))

            if isinstance(it.get('linkTypes'), list):
                for l in it.get('linkTypes'):
                    if l.get('displayName') == 'Children':
                        if isinstance(l.get('linkDTOs'), list):
                            links = "\n\n".join(
                                [ii.get('weburi').split('&id=')[-1]
                                 for ii in l.get('linkDTOs')])
                        elif isinstance(l.get('linkDTOs'), dict):
                            links = (l.get('linkDTOs').get('weburi').split(
                                '&id=')[-1])
                        sheet['BI{}'.format(n)] = "{}".format(links)

                    elif l.get('displayName') == 'Blocks':
                        if isinstance(l.get('linkDTOs'), list):
                            links = "\n\n".join(
                                [ii.get('weburi').split('&id=')[-1]
                                 for ii in l.get('linkDTOs')])
                        elif isinstance(l.get('linkDTOs'), dict):
                            links = (l.get('linkDTOs').get('weburi').split(
                                '&id=')[-1])
                        sheet['BJ{}'.format(n)] = "{}".format(links)

                    elif l.get('displayName') == 'Parents':
                        if isinstance(l.get('linkDTOs'), list):
                            links = "\n\n".join(
                                [ii.get('weburi').split('&id=')[-1]
                                 for ii in l.get('linkDTOs')])
                        elif isinstance(l.get('linkDTOs'), dict):
                            links = (l.get('linkDTOs').get('weburi').split(
                                '&id=')[-1])
                        sheet['BK{}'.format(n)] = "{}".format(links)

            elif isinstance(it.get('linkTypes'), dict):
                l = it.get('linkTypes')
                if l.get('displayName') == 'Children':
                    if isinstance(l.get('linkDTOs'), list):
                        links = "\n\n".join(
                            [ii.get('weburi').split('&id=')[-1]
                             for ii in l.get('linkDTOs')])
                    elif isinstance(l.get('linkDTOs'), dict):
                        links = (l.get('linkDTOs').get('weburi').split(
                            '&id=')[-1])
                    sheet['BI{}'.format(n)] = "{}".format(links)

                elif l.get('displayName') == 'Blocks':
                    if isinstance(l.get('linkDTOs'), list):
                        links = "\n\n".join(
                            [ii.get('weburi').split('&id=')[-1]
                             for ii in l.get('linkDTOs')])
                    elif isinstance(l.get('linkDTOs'), dict):
                        links = (l.get('linkDTOs').get('weburi').split(
                            '&id=')[-1])
                    sheet['BJ{}'.format(n)] = "{}".format(links)

                elif l.get('displayName') == 'Parents':
                    if isinstance(l.get('linkDTOs'), list):
                        links = "\n\n".join(
                            [ii.get('weburi').split('&id=')[-1]
                             for ii in l.get('linkDTOs')])
                    elif isinstance(l.get('linkDTOs'), dict):
                        links = (l.get('linkDTOs').get('weburi').split(
                            '&id=')[-1])
                    sheet['BK{}'.format(n)] = "{}".format(links)

            owner = ''
            domain = ''
            assigned_date = ''
            expiration_date = ''
            rca_date = ''
            closed_date = ''
            closed_week = ''
            assigned_week = ''

            for c in it.get('changes'):
                parsed_history = self.parse_history(c.get('content'))
                sheet_history['A{}'.format(p)] = it.get('itemId')
                sheet_history['B{}'.format(p)] = c.get('modifiedDate')
                sheet_history['C{}'.format(p)] = c.get(
                    'modifiedBy').get('userId')
                sheet_history['D{}'.format(p)] = c.get('content')
                sheet_history['E{}'.format(p)] = json.dumps(parsed_history)
                if (parsed_history.get('Owned By') and
                        "{}&rarr;".format(SPEC_USER) in parsed_history.get('Owned By')):
                    owner = parsed_history.get('Owned By')[15:]
                    sheet_history['F{}'.format(p)] = owner

                    assigned_date = dt.datetime.strptime(
                        c.get('modifiedDate'), "%Y-%m-%dT%H:%M:%S.%fZ")
                    sheet_history['G{}'.format(p)] = assigned_date
                    assigned_week = assigned_date.isocalendar()[1]

                    expiration_date = assigned_date + dt.timedelta(days=7)
                    sheet_history['H{}'.format(p)] = expiration_date

                if parsed_history.get('Root Cause'):
                    rca_date = dt.datetime.strptime(
                        c.get('modifiedDate'), "%Y-%m-%dT%H:%M:%S.%fZ")
                    sheet_history['I{}'.format(p)] = rca_date

                if (parsed_history.get('Tags') and
                        '{}'.format(SPEC_STATUS) in parsed_history.get('Tags')):
                    closed_date = dt.datetime.strptime(
                        c.get('modifiedDate'), "%Y-%m-%dT%H:%M:%S.%fZ")
                    sheet_history['J{}'.format(p)] = closed_date

                    closed_week = closed_date.isocalendar()[1]
                    sheet_history['K{}'.format(p)] = closed_week

                if (parsed_history.get('Filed Against') and
                        '{}&rarr;'.format(PROJECT) in parsed_history.get(
                            'Filed Against')):
                    domain = parsed_history.get('Filed Against')[13:]

                p += 1

            lead_time = dt.timedelta(0)
            if closed_date and assigned_date:
                lead_time = closed_date - assigned_date

            sheet_results['A{}'.format(n)] = int(
                self.get_wi_attr_value(it, 'id')['id'])
            sheet_results['B{}'.format(n)] = self.parse_text(
                self.get_wi_attr_value(it, 'summary')['content'])

            severity = self.get_wi_attr_value(
                it,
                'com.{}.team.workitem.attribute.issueseverity_description')[
                    'label']
            if severity.startswith('1') or severity.startswith('2'):
                sheet_results['C{}'.format(n)] = 'L100'
            else:
                sheet_results['C{}'.format(n)] = 'L50'

            sheet_results['D{}'.format(n)] = self.get_team(owner)
            sheet_results['E{}'.format(n)] = owner
            sheet_results['F{}'.format(n)] = domain
            sheet_results['G{}'.format(n)] = assigned_date
            sheet_results['H{}'.format(n)] = expiration_date
            sheet_results['I{}'.format(n)] = rca_date
            sheet_results['J{}'.format(n)] = closed_date
            sheet_results['K{}'.format(n)] = closed_week
            sheet_results['L{}'.format(n)] = assigned_week
            sheet_results['M{}'.format(n)] = ''  # TODO: RCA
            sheet_results['N{}'.format(n)] = last_comment
            sheet_results['O{}'.format(n)] = self.get_wi_attr_value(
                it, 'internalTags')['content']
            sheet_results['P{}'.format(n)] = lead_time
            sheet_results['Q{}'.format(n)] = self.get_wi_attr_value(
                it, 'internalState')['label']
            sheet_results['R{}'.format(n)] = self.get_wi_attr_value(
                it, 'target')['label']

            rc_field = 'com.{}.team.workitem.attribute.rootcausehtml'.format(CUSTOMER)
            rc_html = self.get_wi_attr_value(it, rc_field)['content']
            sheet_results['S{}'.format(n)] = (
                self.parse_text(rc_html) if rc_html else '')

        wb.save(filepath)
        return True

    @staticmethod
    def parse_history(content):
        html_content = "<html><body>{}</body></html>".format(
            content.encode("utf8"))
        soup = bs.BeautifulSoup(html_content)
        mydict = {}
        for i in soup.body.table:
            key = i.find("td", {"class": "HistoryColumn0"})
            value = i.find("td", {"class": "HistoryColumn1"})
            if key and value and key.text and value.text:
                k = key.text.replace('&nbsp;', '')
                v = value.text.replace('&nbsp;', '')
                mydict[k] = v
            else:
                tds = i.findAll('td')
                if len(tds) > 1:
                    k = tds[0].text.replace('&nbsp;', '')
                    v = tds[1].text.replace('&nbsp;', '')
                else:
                    k = 'None'
                    v = tds[0].text.replace('&nbsp;', '')
                mydict[k] = v
        return mydict

    @staticmethod
    def find_in_content(content, search_key):
        html_content = "<html><body>{}</body></html>".format(
            content.encode("utf8"))
        soup = bs.BeautifulSoup(html_content)
        for i in soup.body.table:
            key = i.find("search_key")
            return key.text

    @staticmethod
    def get_wi_attr_value(wi, key):
        for attr in wi['attributes']:
            if attr['key'] == key:
                return attr['value']

    @staticmethod
    def parse_text(html):
        html_content = "<html><body>{}</body></html>".format(
            html.encode("utf8"))
        soup = bs.BeautifulSoup(html_content,
                                convertEntities=bs.BeautifulSoup.HTML_ENTITIES)
        return ''.join(soup.findAll(text=True))
